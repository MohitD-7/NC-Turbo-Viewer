"""
apply_foutput_priority_6600.py
Apply three updates to catalogue.json:
  Part A: Force re-download for Foutput.csv products
  Part B: Priority fix for ALL cushion products (Wayfair first)
  Part C: 6600-DG and 6701-DG from Jun 08 Furniture Excel
"""

import json
import os
import re
import hashlib
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import openpyxl
from PIL import Image
from io import BytesIO

# ── Paths ──────────────────────────────────────────────────────────────────────
CATALOGUE_PATH = "data/catalogue.json"
THUMB_DIR = "static/thumbnails"
FURNITURE_EXCEL = (
    r"C:\Users\Lenovo\Master Image Library Dropbox\Master Image Library"
    r"\NorthCape Library - Master Excel - Furniture - Jun 08, 2026.xlsx"
)

# ── Part A Part Numbers ────────────────────────────────────────────────────────
FOUTPUT_PART_NUMBERS = [
    "CUSH5619B-CFN", "CUSH5619B-HE", "CUSH5619B-CN", "CUSH5619B-RC",
    "CUSH5619B-RG", "CUSH5619B-RI", "CUSH5619B-RO",
    "CUSH3717B-CFN", "CUSH3717B-HE", "CUSH3717B-CN", "CUSH3717B-RC",
    "CUSH3717B-RG", "CUSH3717B-RI", "CUSH3717B-RO",
    "CUSH4017B-CFN", "CUSH4017B-HE", "CUSH4017B-CN", "CUSH4017B-RC",
    "CUSH4017B-RG", "CUSH4017B-RI", "CUSH4017B-RO",
    "CUSH4818B-CFN", "CUSH4818B-HE", "CUSH4818B-CN", "CUSH4818B-RC",
    "CUSH4818B-RG", "CUSH4818B-RI", "CUSH4818B-RO",
    "CUSH2325DS-CFN", "CUSH2325DS-HE", "CUSH2325DS-CN", "CUSH2325DS-RC",
    "CUSH2325DS-RG", "CUSH2325DS-RI", "CUSH2325DS-RO",
    "CUSH2424DS-CFN", "CUSH2424DS-HE", "CUSH2424DS-CN", "CUSH2424DS-RC",
    "CUSH2424DS-RG", "CUSH2424DS-RI", "CUSH2424DS-RO",
    "CUSH2526DS-CFN", "CUSH2526DS-HE", "CUSH2526DS-CN", "CUSH2526DS-RC",
    "CUSH2526DS-RG", "CUSH2526DS-RI", "CUSH2526DS-RO",
    "CUSH2630DS-CFN", "CUSH2630DS-HE", "CUSH2630DS-CN", "CUSH2630DS-RC",
    "CUSH2630DS-RG", "CUSH2630DS-RI", "CUSH2630DS-RO",
    "CUSH271SCC-CFN", "CUSH271SCC-HE", "CUSH271SCC-CN", "CUSH271SCC-RC",
    "CUSH271SCC-RG", "CUSH271SCC-RI", "CUSH271SCC-RO",
    "CUSH2178CL-CFN", "CUSH2178CL-HE", "CUSH2178CL-CN", "CUSH2178CL-RC",
    "CUSH2178CL-RG", "CUSH2178CL-RI", "CUSH2178CL-RO",
    "CUSH2375CL-CFN", "CUSH2375CL-HE", "CUSH2375CL-CN", "CUSH2375CL-RC",
    "CUSH2375CL-RG", "CUSH2375CL-RI", "CUSH2375CL-RO",
    "CUSH2473CL-CFN", "CUSH2473CL-HE", "CUSH2473CL-CN", "CUSH2473CL-RC",
    "CUSH2473CL-RG", "CUSH2473CL-RI", "CUSH2473CL-RO",
    "CUSH2680CL-CFN", "CUSH2680CL-HE", "CUSH2680CL-CN", "CUSH2680CL-RC",
    "CUSH2680CL-RG", "CUSH2680CL-RI", "CUSH2680CL-RO",
    "CUSH2022O-CFN", "CUSH2022O-HE", "CUSH2022O-CN", "CUSH2022O-RC",
    "CUSH2022O-RG", "CUSH2022O-RI", "CUSH2022O-RO",
    "CUSH2319O-CFN", "CUSH2319O-HE", "CUSH2319O-CN", "CUSH2319O-RC",
    "CUSH2319O-RG", "CUSH2319O-RI", "CUSH2319O-RO",
    "CUSH2424O-CFN", "CUSH2424O-HE", "CUSH2424O-CN", "CUSH2424O-RC",
    "CUSH2424O-RG", "CUSH2424O-RI", "CUSH2424O-RO",
    "CUSH2624O-CFN", "CUSH2624O-HE", "CUSH2624O-CN", "CUSH2624O-RC",
    "CUSH2624O-RG", "CUSH2624O-RI", "CUSH2624O-RO",
    # 1616DC-CFN treated as CUSH1616DC-CFN — but in catalogue it's stored as "1616DC-CFN"
    # We'll handle BOTH name variants
    "CUSH1616DC-CFN", "CUSH1616DC-HE", "CUSH1616DC-CN", "CUSH1616DC-RC",
    "CUSH1616DC-RG", "CUSH1616DC-RI", "CUSH1616DC-RO",
    "CUSH1818DC-CFN", "CUSH1818DC-HE", "CUSH1818DC-CN", "CUSH1818DC-RC",
    "CUSH1818DC-RG", "CUSH1818DC-RI", "CUSH1818DC-RO",
    "CUSH2020DC-CFN", "CUSH2020DC-HE", "CUSH2020DC-CN", "CUSH2020DC-RC",
    "CUSH2020DC-RG", "CUSH2020DC-RI", "CUSH2020DC-RO",
    "CUSH2119DC-CFN", "CUSH2119DC-HE", "CUSH2119DC-CN", "CUSH2119DC-RC",
    "CUSH2119DC-RG", "CUSH2119DC-RI", "CUSH2119DC-RO",
    # Also try without CUSH prefix for the 1616DC set (as noted in spec)
    "1616DC-CFN", "1616DC-HE", "1616DC-CN", "1616DC-RC",
    "1616DC-RG", "1616DC-RI", "1616DC-RO",
]
# Deduplicate while preserving order
seen = set()
FOUTPUT_PART_NUMBERS_DEDUP = []
for pn in FOUTPUT_PART_NUMBERS:
    if pn not in seen:
        seen.add(pn)
        FOUTPUT_PART_NUMBERS_DEDUP.append(pn)


# ── Helpers ────────────────────────────────────────────────────────────────────

def url_to_thumb_filename(url: str) -> str:
    """md5(url.encode()).hexdigest() + '.jpg'"""
    return hashlib.md5(url.encode()).hexdigest() + ".jpg"


def url_to_download_url(url: str) -> str:
    """Convert Dropbox share URL to direct download URL."""
    u = url.replace("dl=0", "raw=1").replace("dl=1", "raw=1")
    if "raw=1" not in u:
        sep = "&" if "?" in u else "?"
        u = u + sep + "raw=1"
    return u


def download_thumb(url: str, thumb_dir: str) -> str | None:
    """
    Download URL, resize to 400x400 JPEG q85, save to thumb_dir.
    Returns relative path 'thumbnails/<md5>.jpg' on success, None on failure.
    """
    filename = url_to_thumb_filename(url)
    dest = os.path.join(thumb_dir, filename)
    if os.path.exists(dest):
        return f"thumbnails/{filename}"
    dl_url = url_to_download_url(url)
    try:
        resp = requests.get(dl_url, timeout=30)
        resp.raise_for_status()
        img = Image.open(BytesIO(resp.content)).convert("RGB")
        img = img.resize((400, 400), Image.LANCZOS)
        img.save(dest, "JPEG", quality=85)
        return f"thumbnails/{filename}"
    except Exception as e:
        print(f"  FAIL download {url[:60]}: {e}")
        return None


def get_url_list(product: dict, source: str) -> list[str]:
    """Get all non-empty URLs for a given source from product dict."""
    if source == "wf":
        keys = [f"Wayfair Image {i}" for i in range(1, 16)]
    elif source == "nc":
        keys = [f"Northcape Image {i}" for i in range(1, 16)]
    elif source == "by":
        keys = (
            [f"Overstock Image {i}" for i in range(1, 16)]
            + [f"Bed Bath Beyond Image {i}" for i in range(1, 16)]
        )
    elif source == "hd":
        keys = [f"Home Depot Image {i}" for i in range(1, 16)]
    else:
        return []
    return [product[k] for k in keys if product.get(k)]


def build_priority_url_list(product: dict) -> list[str]:
    """
    Build ordered URL list: Wayfair first (up to 5), then NC, then BY, then HD.
    Fill to max 5 total.
    """
    urls = []
    for source in ("wf", "nc", "by", "hd"):
        for url in get_url_list(product, source):
            if url not in urls:
                urls.append(url)
            if len(urls) >= 5:
                return urls
    return urls


def delete_thumb(filename: str, thumb_dir: str) -> bool:
    """Delete a thumbnail file. filename is like 'thumbnails/abc.jpg' or just 'abc.jpg'."""
    basename = os.path.basename(filename)
    path = os.path.join(thumb_dir, basename)
    if os.path.exists(path):
        os.remove(path)
        return True
    return False


def rebuild_image_list(
    product: dict,
    thumb_dir: str,
    executor: ThreadPoolExecutor,
    download_counter: list,
    fail_counter: list,
) -> tuple[list[str], str]:
    """
    Download thumbnails for product using priority order.
    Returns (new_image_list, new_local_thumbnail).
    """
    priority_urls = build_priority_url_list(product)
    futures = {executor.submit(download_thumb, url, thumb_dir): url for url in priority_urls}
    results_by_url = {}
    for future in as_completed(futures):
        url = futures[future]
        path = future.result()
        results_by_url[url] = path
        if path:
            download_counter[0] += 1
        else:
            fail_counter[0] += 1
        if download_counter[0] % 50 == 0:
            print(f"  ... {download_counter[0]} thumbnails downloaded so far")

    new_image_list = []
    for url in priority_urls:
        p = results_by_url.get(url)
        if p:
            new_image_list.append(p)

    local_thumb = new_image_list[0] if new_image_list else ""
    return new_image_list, local_thumb


def get_hyperlink_parts(cell_value: str) -> tuple[str | None, str | None]:
    """
    Parse =HYPERLINK("url","display") formula.
    Returns (url, display_text).
    """
    if not cell_value or not isinstance(cell_value, str):
        return cell_value, None
    if cell_value.startswith("=HYPERLINK("):
        m = re.match(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', cell_value)
        if m:
            return m.group(1), m.group(2)
        m2 = re.match(r'=HYPERLINK\("([^"]+)"\)', cell_value)
        if m2:
            return m2.group(1), None
    return cell_value, None


# ── Load catalogue ─────────────────────────────────────────────────────────────

print("Loading catalogue...")
with open(CATALOGUE_PATH, "r", encoding="utf-8") as f:
    catalogue = json.load(f)

# Index by Part Number for fast lookup
pn_index = {p["Part Number"]: i for i, p in enumerate(catalogue)}

os.makedirs(THUMB_DIR, exist_ok=True)


# ── Part A: Force re-download for Foutput.csv products ────────────────────────
print("\n=== Part A: Force re-download for Foutput.csv products ===")

partA_products = 0
partA_deleted = 0
partA_downloaded = 0
partA_failed = 0

foutput_set = set(FOUTPUT_PART_NUMBERS_DEDUP)

# Find matching catalogue entries
partA_targets = []
for pn in FOUTPUT_PART_NUMBERS_DEDUP:
    idx = pn_index.get(pn)
    if idx is not None:
        partA_targets.append((pn, idx))
    else:
        print(f"  NOT IN CATALOGUE: {pn}")

print(f"Found {len(partA_targets)} products in catalogue for Part A")

dl_counter = [0]
fail_counter = [0]

with ThreadPoolExecutor(max_workers=8) as executor:
    for pn, idx in partA_targets:
        product = catalogue[idx]
        old_image_list = product.get("Image_List", [])

        # Delete existing thumbnails
        deleted = 0
        for img_path in old_image_list:
            if delete_thumb(img_path, THUMB_DIR):
                deleted += 1
        partA_deleted += deleted

        # Re-download with priority order
        priority_urls = build_priority_url_list(product)
        new_image_list = []
        futures = {
            executor.submit(download_thumb, url, THUMB_DIR): url
            for url in priority_urls
        }
        results_by_url = {}
        for future in as_completed(futures):
            url = futures[future]
            path = future.result()
            results_by_url[url] = path
            if path:
                dl_counter[0] += 1
                partA_downloaded += 1
            else:
                fail_counter[0] += 1
                partA_failed += 1
            if dl_counter[0] % 50 == 0:
                print(f"  ... {dl_counter[0]} thumbnails downloaded so far")

        for url in priority_urls:
            p = results_by_url.get(url)
            if p:
                new_image_list.append(p)

        product["Image_List"] = new_image_list
        product["Local_Thumbnail"] = new_image_list[0] if new_image_list else ""
        partA_products += 1

print(f"Part A done: {partA_products} products, {partA_deleted} deleted, {partA_downloaded} downloaded, {partA_failed} failed")


# ── Part B: Priority fix for ALL cushion products ─────────────────────────────
print("\n=== Part B: Priority fix for all cushion products ===")

partB_products = 0
partB_swapped = 0
partB_failed = 0

# All cushion products (Category=Cushion or Cushions, or Type contains cushion)
def is_cushion(product: dict) -> bool:
    cat = (product.get("Category") or "").lower()
    typ = (product.get("Type") or "").lower()
    return "cushion" in cat or "cushion" in typ

# Already processed in Part A — skip those
partA_pns = {pn for pn, _ in partA_targets}

cushion_products = [
    (i, p)
    for i, p in enumerate(catalogue)
    if is_cushion(p) and p.get("Part Number") not in partA_pns
]
print(f"Checking {len(cushion_products)} cushion products (excluding Part A)...")

partB_dl_counter = [0]
partB_fail_counter = [0]

with ThreadPoolExecutor(max_workers=8) as executor:
    for idx, product in cushion_products:
        wf_urls = get_url_list(product, "wf")
        nc_urls = get_url_list(product, "nc")
        by_urls = get_url_list(product, "by")

        if not wf_urls:
            # No Wayfair images — nothing to prioritize
            continue

        current_image_list = product.get("Image_List", [])
        if not current_image_list:
            continue

        # Compute md5 filenames for Wayfair URLs
        wf_filenames = set(url_to_thumb_filename(u) for u in wf_urls)
        # Compute md5 filenames for current Image_List
        current_filenames = set(os.path.basename(p) for p in current_image_list)

        # Check if ANY Wayfair image is missing from current list
        # AND current list has non-WF images (NC or BY)
        nc_filenames = set(url_to_thumb_filename(u) for u in nc_urls)
        by_filenames = set(url_to_thumb_filename(u) for u in by_urls)

        has_non_wf = bool(current_filenames & (nc_filenames | by_filenames))
        missing_wf = bool(wf_filenames - current_filenames)

        if not (has_non_wf and missing_wf):
            # Either already all-WF, or WF is already fully represented
            # But also check if first 5 slots aren't WF-first
            # Build what the ideal list would be
            ideal_priority = build_priority_url_list(product)
            ideal_filenames = [url_to_thumb_filename(u) for u in ideal_priority]
            if ideal_filenames == [os.path.basename(p) for p in current_image_list]:
                continue  # Already correct
            # Also skip if no Wayfair images would replace anything
            if not missing_wf:
                continue

        # Rebuild: delete non-WF thumbnails that WF could replace
        # Delete all current thumbnails
        swapped = 0
        for img_path in current_image_list:
            fname = os.path.basename(img_path)
            if fname in (nc_filenames | by_filenames) and fname not in wf_filenames:
                if delete_thumb(img_path, THUMB_DIR):
                    swapped += 1

        # Download missing WF thumbnails and rebuild list
        priority_urls = build_priority_url_list(product)
        futures = {
            executor.submit(download_thumb, url, THUMB_DIR): url
            for url in priority_urls
        }
        results_by_url = {}
        for future in as_completed(futures):
            url = futures[future]
            path = future.result()
            results_by_url[url] = path
            if path:
                partB_dl_counter[0] += 1
            else:
                partB_fail_counter[0] += 1
                partB_failed += 1
            if (partB_dl_counter[0] + dl_counter[0]) % 50 == 0:
                total = dl_counter[0] + partB_dl_counter[0]
                print(f"  ... {total} total thumbnails downloaded so far")

        new_image_list = []
        for url in priority_urls:
            p = results_by_url.get(url)
            if p:
                new_image_list.append(p)

        if new_image_list != current_image_list:
            product["Image_List"] = new_image_list
            product["Local_Thumbnail"] = new_image_list[0] if new_image_list else ""
            partB_products += 1
            partB_swapped += swapped

print(f"Part B done: {partB_products} products priority-fixed, {partB_swapped} thumbnails swapped, {partB_failed} failed")


# ── Part C: 6600-DG and 6701-DG from Jun 08 Furniture Excel ─────────────────
print("\n=== Part C: 6600-DG and 6701-DG from Furniture Excel ===")

wb = openpyxl.load_workbook(FURNITURE_EXCEL)

partC_added = 0
partC_updated = 0
partC_downloaded = 0
partC_failed = 0

IMAGE_COL_NAMES = (
    [f"Northcape Image {i}" for i in range(1, 16)]
    + [f"Overstock Image {i}" for i in range(1, 16)]
    + [f"Wayfair Image {i}" for i in range(1, 16)]
    + [f"Home Depot Image {i}" for i in range(1, 16)]
    + [f"Bed Bath Beyond Image {i}" for i in range(1, 16)]
)
# Other fields to copy from Excel
OTHER_FIELDS = [
    "Dropbox Folder Path", "Type", "Collection", "Arm/Table-Top", "Product",
    "Last Modified", "NC Image Count", "BY Image Count", "WF Image Count", "HD Image Count",
    "Collection Type",
]

def read_sheet_products(ws) -> list[dict]:
    """Read all data rows from a sheet, return list of dicts."""
    headers = [c.value for c in ws[1]]
    products = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_dict = {}
        for cell in row:
            h = headers[cell.column - 1] if cell.column - 1 < len(headers) else None
            if h:
                row_dict[h] = cell.value
        pn = row_dict.get("Part Number")
        if not pn:
            continue
        # Parse Color hyperlink
        color_raw = row_dict.get("Color")
        color_url, color_display = get_hyperlink_parts(str(color_raw) if color_raw else "")
        row_dict["Color"] = color_display or color_raw
        row_dict["Color_Link"] = color_url if color_url != color_raw else ""
        products.append(row_dict)
    return products


def upsert_product(
    excel_row: dict,
    catalogue: list,
    pn_index: dict,
    thumb_dir: str,
    executor: ThreadPoolExecutor,
    dl_counter: list,
    fail_counter: list,
) -> tuple[str, int]:
    """
    Insert or update a product.
    Excel Part Numbers are NCI-prefixed; catalogue uses NC-prefix.
    Returns ("added"/"updated", thumbnails_downloaded).
    """
    excel_pn = excel_row.get("Part Number", "")
    # Map NCI -> NC for catalogue
    cat_pn = excel_pn.replace("NCI", "NC", 1) if excel_pn.startswith("NCI") else excel_pn

    # Build image URL fields from excel row
    image_fields = {}
    for col in IMAGE_COL_NAMES:
        v = excel_row.get(col)
        if v and isinstance(v, str) and v.strip():
            image_fields[col] = v.strip()

    # Priority URL list for this excel row
    # Treat excel_row as a product dict for get_url_list
    priority_urls = []
    for source in ("wf", "nc", "by", "hd"):
        for url in get_url_list(excel_row, source):
            if url not in priority_urls:
                priority_urls.append(url)
            if len(priority_urls) >= 5:
                break
        if len(priority_urls) >= 5:
            break

    # Download thumbnails
    futures = {
        executor.submit(download_thumb, url, thumb_dir): url
        for url in priority_urls
    }
    results_by_url = {}
    thumbs_dl = 0
    for future in as_completed(futures):
        url = futures[future]
        path = future.result()
        results_by_url[url] = path
        if path:
            dl_counter[0] += 1
            thumbs_dl += 1
        else:
            fail_counter[0] += 1
        if dl_counter[0] % 50 == 0:
            print(f"  ... {dl_counter[0]} Part C thumbnails downloaded so far")

    new_image_list = []
    for url in priority_urls:
        p = results_by_url.get(url)
        if p:
            new_image_list.append(p)
    local_thumb = new_image_list[0] if new_image_list else ""

    if cat_pn in pn_index:
        # Update existing
        idx = pn_index[cat_pn]
        product = catalogue[idx]
        # Update image URL fields
        for col, val in image_fields.items():
            product[col] = val
        # Update other fields
        for field in OTHER_FIELDS:
            if excel_row.get(field) is not None:
                product[field] = excel_row[field]
        product["Color"] = excel_row.get("Color", product.get("Color", ""))
        product["Color_Link"] = excel_row.get("Color_Link", product.get("Color_Link", ""))
        product["Image_List"] = new_image_list
        product["Local_Thumbnail"] = local_thumb
        return "updated", thumbs_dl
    else:
        # Add new product
        # Derive Collection Type from part number
        m = re.search(r"NC(?:I?)(\d+)", excel_pn)
        collection_type = m.group(1) if m else ""

        new_product = {
            "Collection Type": collection_type,
            "Thumbnail": "",
            "Dropbox Folder Path": excel_row.get("Dropbox Folder Path", ""),
            "Part Number": cat_pn,
            "Category": "Furniture",
            "Type": excel_row.get("Type", "Furniture"),
            "Collection": excel_row.get("Collection", ""),
            "Color": excel_row.get("Color", ""),
            "Color_Link": excel_row.get("Color_Link", ""),
            "Last Modified": excel_row.get("Last Modified", ""),
            "NC Image Count": excel_row.get("NC Image Count", ""),
            "BY Image Count": excel_row.get("BY Image Count", ""),
            "WF Image Count": excel_row.get("WF Image Count", ""),
            "HD Image Count": excel_row.get("HD Image Count", ""),
            "Arm/Table-Top": excel_row.get("Arm/Table-Top", ""),
            "Product": excel_row.get("Product", ""),
            "_thumbnail_path": "",
            "Image_List": new_image_list,
            "Local_Thumbnail": local_thumb,
        }
        # Add image URL fields
        for col, val in image_fields.items():
            new_product[col] = val

        catalogue.append(new_product)
        pn_index[cat_pn] = len(catalogue) - 1
        return "added", thumbs_dl


partC_dl_counter = [0]
partC_fail_counter = [0]

with ThreadPoolExecutor(max_workers=8) as executor:
    for sheet_name in ["6600 - DG", "6701 - DG"]:
        ws = wb[sheet_name]
        sheet_rows = read_sheet_products(ws)
        print(f"  Sheet '{sheet_name}': {len(sheet_rows)} products")
        for row in sheet_rows:
            action, thumbs = upsert_product(
                row, catalogue, pn_index, THUMB_DIR,
                executor, partC_dl_counter, partC_fail_counter,
            )
            if action == "added":
                partC_added += 1
            else:
                partC_updated += 1
            partC_downloaded += thumbs

partC_failed = partC_fail_counter[0]
print(f"Part C done: {partC_added} added, {partC_updated} updated, {partC_downloaded} thumbnails downloaded, {partC_failed} failed")

# ── Save catalogue ─────────────────────────────────────────────────────────────
print("\nSaving catalogue.json...")
with open(CATALOGUE_PATH, "w", encoding="utf-8") as f:
    json.dump(catalogue, f, ensure_ascii=False, indent=2)
print("Saved.")

# ── Summary ────────────────────────────────────────────────────────────────────
total_failures = partA_failed + partB_failed + partC_failed
print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
print(f"Part A: {partA_products} products force-redownloaded, {partA_deleted} thumbnails deleted, {partA_downloaded} thumbnails downloaded")
print(f"Part B: {partB_products} products priority-fixed, {partB_swapped} thumbnails swapped")
print(f"Part C: {partC_added} new 6600/6701 products added, {partC_updated} updated, {partC_downloaded} thumbnails downloaded")
print(f"Failures: {total_failures}")
