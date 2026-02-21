import streamlit as st
import json
import os
import pandas as pd
import base64
import streamlit.components.v1 as components

# Page Configuration
st.set_page_config(
    page_title="NorthCape Turbo Catalogue",
    page_icon="🛋️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Force Refresh Commit: Triggering Deployment Rebuild
# Custom Styling
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
    
    /* White-labeling: Hide Streamlit Branding & Menus */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    [data-testid="stToolbar"] {visibility: hidden !important;}
    [data-testid="stSidebarCollapseButton"] {display: none !important;}
    
    .stApp {
        background-color: #f8fafc;
        font-family: 'Inter', sans-serif;
    }

    /* Reduce default Streamlit padding */
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 0rem !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    
    /* Container for the responsive grid */
    .card-grid {
        display: grid;
        grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
        gap: 24px;
        width: 100%;
        padding-top: 1.5rem;
    }
    
    /* Strictly 3-5 columns on Desktop */
    @media (min-width: 900px) and (max-width: 1199px) {
        .card-grid { grid-template-columns: repeat(3, 1fr); }
    }
    @media (min-width: 1200px) and (max-width: 1599px) {
        .card-grid { grid-template-columns: repeat(4, 1fr); }
    }
    @media (min-width: 1600px) {
        .card-grid { grid-template-columns: repeat(5, 1fr); }
    }

    /* Product Card */
    .product-card {
        background: white;
        padding: 0.75rem;
        border-radius: 14px;
        border: 1px solid #e2e8f0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.04);
        transition: all 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.1);
        display: flex;
        flex-direction: column;
        height: 100%;
        max-width: 320px;
    }
    
    .product-card:hover {
        transform: translateY(-8px);
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        border-color: #3b82f6;
    }
    
    .image-container {
        width: 100%;
        background: white;
        border-radius: 12px;
        aspect-ratio: 1.2 / 1;
        display: flex;
        align-items: center;
        justify-content: center;
        overflow: hidden;
        margin: 0.5rem 0;
    }
    
    .image-container img {
        width: 100%;
        height: 100%;
        object-fit: contain;
        transition: transform 0.5s ease;
    }
    
    .product-card:hover .image-container img {
        transform: scale(1.25);
    }
    
    /* Swap Image Button */
    .image-container {
        position: relative;
    }
    
    .swap-btn {
        position: absolute;
        bottom: 8px;
        right: 8px;
        background: rgba(255, 255, 255, 0.9);
        border: 1px solid #e2e8f0;
        border-radius: 50%;
        width: 32px;
        height: 32px;
        display: flex;
        align-items: center;
        justify-content: center;
        cursor: pointer;
        font-size: 1.1rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        z-index: 10;
        transition: all 0.2s;
        opacity: 0.6;
    }
    
    .swap-btn:hover {
        opacity: 1;
        background: white;
        transform: scale(1.1);
        border-color: #3b82f6;
    }
    
    .product-card:hover .swap-btn {
        opacity: 0.9;
    }
    
    /* Shortlist Button */
    .shortlist-btn {
        position: absolute;
        top: 6px;
        right: 6px;
        background: rgba(255, 255, 255, 0.95);
        border: 1px solid #e2e8f0;
        border-radius: 50%;
        width: 34px;
        height: 34px;
        display: flex;
        align-items: center;
        justify-content: center;
        cursor: pointer;
        font-size: 1.2rem;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
        z-index: 999;
        transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
        opacity: 0.8;
        pointer-events: auto !important;
    }
    
    .shortlist-btn:hover {
        transform: scale(1.1);
        background: #fff;
        opacity: 1;
        box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1);
    }
    
    .shortlist-btn:active {
        transform: scale(0.9);
        background: #f1f5f9;
    }
    
    .shortlist-btn.active {
        opacity: 1 !important;
        color: #eab308; /* Yellow-500 */
        background: white;
        border-color: #eab308;
    }
    
    .product-card:hover .shortlist-btn {
        opacity: 0.7;
    }
    
    .shortlist-btn:hover {
        opacity: 1 !important;
        transform: scale(1.1);
    }
    
    .card-header {
        display: flex;
        flex-direction: column;
    }
    
    .badge {
        background: #dbeafe;
        color: #1e40af;
        padding: 0.2rem 0.6rem;
        border-radius: 6px;
        font-size: 0.7rem;
        font-weight: 700;
        display: inline-block;
        width: fit-content;
        margin-bottom: 0.5rem;
    }
    
    .part-number {
        font-size: 1.05rem;
        font-weight: 800;
        color: #0f172a;
        margin-bottom: 1px;
    }
    
    .collection-text {
        font-size: 0.85rem;
        color: #64748b;
        margin-bottom: 0.2rem;
    }
    
    .card-footer {
        margin-top: auto;
    }
    
    .detail-row {
        display: flex;
        justify-content: space-between;
        font-size: 0.8rem;
        padding: 0.2rem 0;
    }
    
    .detail-label {
        color: #64748b;
        font-weight: 500;
    }
    
    .detail-value {
        color: #0f172a;
        font-weight: 700;
        text-align: right;
    }
    
    .color-link {
        color: #2563eb;
        text-decoration: none;
    }
    
    .color-link:hover {
        text-decoration: underline;
    }
    
    .color-link {
        color: #2563eb;
        text-decoration: none;
        transition: color 0.2s;
        font-weight: 600;
    }
    
    .color-link:hover {
        color: #1d4ed8;
        text-decoration: underline;
    }
    
    /* Custom Scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
    }
    ::-webkit-scrollbar-track {
        background: #f1f5f9;
    }
    ::-webkit-scrollbar-thumb {
        background: #cbd5e1;
        border-radius: 10px;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: #94a3b8;
    }

    /* Highlighted Search Bar (Match Reference) */
    div[data-baseweb="input"] {
        border: 1.5px solid #bfdbfe !important;
        border-radius: 12px !important;
        transition: all 0.3s;
        background: white !important;
        padding-left: 8px !important;
    }
    div[data-baseweb="input"]:focus-within {
        border-color: #2563eb !important;
        box-shadow: 0 0 0 1px #2563eb !important;
    }
    
    /* Sidebar Label Styling (Match Reference) */
    [data-testid="stWidgetLabel"] p {
        color: #1e40af !important;
        font-weight: 800 !important;
        text-transform: uppercase !important;
        font-size: 0.75rem !important;
        letter-spacing: 0.05em !important;
    }
    
    /* Premium Header */
    .hero-container {
        padding: 1rem 0rem;
        margin-bottom: 0rem;
    }
    
    .hero-title {
        font-size: 2.75rem;
        font-weight: 800;
        color: #1e40af;
        letter-spacing: -0.01em;
        margin-bottom: 0.75rem;
    }
    
    /* Invisible Sync Input - Positioned far off screen but still 'available' to DOM */
    div[data-testid="stTextInput"]:has(input[placeholder="sync_bridge_v7"]) {
        position: fixed;
        left: -100vw;
        top: -100vh;
        z-index: -9999;
        opacity: 0;
        pointer-events: none;
    }
</style>
""", unsafe_allow_html=True)

# Data Loading with Caching
@st.cache_data
def load_catalogue_data():
    # Use absolute path relative to this script's directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(base_dir, "data", "catalogue.json")
    
    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)
    print(f"Warning: Data file not found at {json_path}")
    return []

data = load_catalogue_data()

# --- Shortlist Session State ---
if 'shortlist' not in st.session_state:
    st.session_state.shortlist = set() # Store Part Numbers for uniqueness

if 'view_shortlist' not in st.session_state:
    st.session_state.view_shortlist = False
if "sync_counter" not in st.session_state:
    st.session_state.sync_counter = 0

# Helper for Base64 Thumbnails (Fixes all Cloud/Local pathing issues)
def get_base64_img(thumb_path):
    if not thumb_path: return None
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        # Handle cases where path might be 'thumbnails/foo.jpg' or just 'foo.jpg'
        fname = os.path.basename(thumb_path)
        abs_path = os.path.join(base_dir, "static", "thumbnails", fname)
        
        if os.path.exists(abs_path):
            with open(abs_path, "rb") as f:
                data = base64.b64encode(f.read()).decode()
                return f"data:image/jpeg;base64,{data}"
    except Exception:
        pass
    return None

# --- Main UI Rendering (Optimized with st.fragment) ---
@st.fragment
def render_main_ui(filtered_df):
    # 1. Shortlist Sync Bridge (V7 - Local Re-run)
    sync_key = f"sync_v7_{st.session_state.sync_counter}"
    # Invisible input for JS communication
    sync_val = st.text_input("sync_bridge", placeholder="sync_bridge_v7", key=sync_key, label_visibility="collapsed")
    
    if sync_val and "|" in sync_val:
        try:
            part = sync_val.split("|")[0]
            if part in st.session_state.shortlist:
                st.session_state.shortlist.remove(part)
                st.toast(f"Removed: {part}", icon="🗑️")
            else:
                st.session_state.shortlist.add(part)
                st.toast(f"Added: {part}", icon="⭐")
            st.session_state.sync_counter += 1
            st.rerun(scope="fragment")
        except Exception:
            pass

    # 2. Sidebar Shortlist & Export Section (In Fragment to update count instantly)
    with st.sidebar:
        st.divider()
        st.markdown(f"### ⭐ Shortlist ({len(st.session_state.shortlist)})")
        
        # View Shortlist Only Toggle
        view_mode = st.toggle("View Shortlist Only", value=st.session_state.view_shortlist)
        if view_mode != st.session_state.view_shortlist:
            st.session_state.view_shortlist = view_mode
            st.rerun() # Full rerun needed to update global filtering

        if st.session_state.view_shortlist:
            filtered_df = filtered_df[filtered_df["Part Number"].isin(st.session_state.shortlist)]

        # Shortlist All Visible Button
        if not filtered_df.empty:
            if st.button("Shortlist All Visible", use_container_width=True):
                visible_parts = set(filtered_df["Part Number"].astype(str).tolist())
                st.session_state.shortlist.update(visible_parts)
                st.rerun(scope="fragment")

        # Clear Shortlist Button
        if st.button("Clear All", use_container_width=True):
            st.session_state.shortlist = set()
            st.rerun(scope="fragment")

        # --- Export Section ---
        if len(st.session_state.shortlist) > 0:
            st.divider()
            st.markdown("### 📥 Export Shortlist")
            export_format = st.selectbox("Choose Format", ["Excel (.xlsx)", "PDF Gallery"])
            
            # Use all data for export, not just filtered subset
            export_raw_df = df[df["Part Number"].isin(st.session_state.shortlist)]
            
            if export_format == "Excel (.xlsx)":
                try:
                    import io
                    import glob
                    import openpyxl.utils
                    from openpyxl.styles import Font
                    output = io.BytesIO()
                    base_dir = os.path.dirname(os.path.abspath(__file__))
                    excel_files = glob.glob(os.path.join(base_dir, "*.xlsx"))
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        unique_types = export_raw_df["Collection Type"].unique()
                        for coll_type in unique_types:
                            parts_subset = export_raw_df[export_raw_df["Collection Type"] == coll_type]
                            parts_list = parts_subset["Part Number"].tolist()
                            
                            source_df = None
                            for f in excel_files:
                                try:
                                    xl = pd.ExcelFile(f)
                                    if coll_type in xl.sheet_names:
                                        source_df = pd.read_excel(f, sheet_name=coll_type)
                                        source_df = source_df[source_df["Part Number"].isin(parts_list)]
                                        break
                                except Exception: continue
                            
                            final_sheet_df = source_df.copy() if source_df is not None else parts_subset.copy()
                            drop_cols = ["Thumbnail", "_thumbnail_path", "Local_Thumbnail", "Image_List", "Part Number_Link", "Collection Type"]
                            final_sheet_df = final_sheet_df.drop(columns=[c for c in drop_cols if c in final_sheet_df.columns])
                            
                            if "Color_Link" not in final_sheet_df.columns:
                                link_map = df[["Part Number", "Color_Link"]].drop_duplicates()
                                final_sheet_df = final_sheet_df.merge(link_map, on="Part Number", how="left")
                            
                            # Ordering Logic
                            cols = final_sheet_df.columns.tolist()
                            if "Part Number" in cols: cols.insert(0, cols.pop(cols.index("Part Number")))
                            if "Product" in cols:
                                p_idx = cols.index("Product")
                                if "Arm/Table-Top" in cols:
                                    cols.insert(p_idx, cols.pop(cols.index("Arm/Table-Top")))
                                    p_idx = cols.index("Product")
                                if "Panel" in cols: cols.insert(p_idx + 1, cols.pop(cols.index("Panel")))
                            
                            final_sheet_df = final_sheet_df[cols]
                            sheet_name = "".join([c for c in str(coll_type) if c not in r'[]:*?/\ '])[:31]
                            final_sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
                            
                            worksheet = writer.sheets[sheet_name]
                            for col in worksheet.columns:
                                max_length = max([len(str(cell.value or "")) for cell in col])
                                worksheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)

                            color_col = "Cushion Color" if coll_type == "Cushions" else "Color"
                            h_font = Font(size=11, underline='single', color='0563C1')
                            
                            if color_col in final_sheet_df.columns and "Color_Link" in final_sheet_df.columns:
                                c_idx = final_sheet_df.columns.get_loc(color_col) + 1
                                l_idx = final_sheet_df.columns.get_loc("Color_Link") + 1
                                for row_num in range(2, len(final_sheet_df) + 2):
                                    link_val = worksheet.cell(row=row_num, column=l_idx).value
                                    if link_val and str(link_val).startswith("http"):
                                        cell = worksheet.cell(row=row_num, column=c_idx)
                                        cell.hyperlink = link_val
                                        cell.font = h_font
                                worksheet.delete_cols(l_idx)
                                
                            current_headers = [cell.value for cell in worksheet[1]]
                            if "Dropbox Folder Path" in current_headers:
                                p_idx_ws = current_headers.index("Dropbox Folder Path") + 1
                                worksheet.column_dimensions[openpyxl.utils.get_column_letter(p_idx_ws)].visible = False
                    
                    st.download_button("Download Excel", data=output.getvalue(), file_name="NC_Shortlist.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                except Exception as e:
                    st.error(f"Excel Export failed: {str(e)}")

            elif export_format == "PDF Gallery":
                try:
                    from fpdf import FPDF
                    class PDF(FPDF):
                        def header(self):
                            self.set_font('helvetica', 'B', 22); self.set_text_color(30, 64, 175); self.cell(0, 15, 'NORTHCAPE CATALOGUE', 0, 1, 'C')
                            self.set_draw_color(226, 232, 240); self.line(10, 25, 200, 25); self.ln(10)
                        def footer(self): self.set_y(-15); self.set_font('helvetica', 'I', 8); self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
                    
                    pdf = PDF(); pdf.set_auto_page_break(auto=False, margin=0); pdf.add_page()
                    margin, gutter, col_width, row_height = 10, 5, 60, 85
                    current_col, current_row = 0, 0
                    
                    for i, (_, item) in enumerate(export_raw_df.iterrows()):
                        if i > 0 and i % 9 == 0:
                            pdf.add_page(); current_col, current_row = 0, 0
                        cell_x, cell_y = margin + (current_col * (col_width + gutter)), 30 + (current_row * row_height)
                        thumb_path = item.get('Local_Thumbnail')
                        img_y_offset = cell_y
                        if thumb_path:
                            abs_thumb = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "thumbnails", os.path.basename(thumb_path))
                            if os.path.exists(abs_thumb):
                                img_w = 58; pdf.image(abs_thumb, x=cell_x + (col_width - img_w) / 2, y=cell_y, w=img_w)
                                img_y_offset += 48
                        
                        pdf.set_xy(cell_x, img_y_offset + 2); pdf.set_font('helvetica', 'B', 8); pdf.set_text_color(15, 23, 42)
                        pdf.multi_cell(col_width, 4, str(item['Part Number']), ln=0, align='C')
                        
                        pdf.set_font('helvetica', '', 7); details_y = pdf.get_y() + 1
                        st_type, prod_l = str(item.get('Type', '')).strip(), str(item.get('Product', '')).lower()
                        fields = [("Type", item.get('Type')), ("Collection", item.get('Collection'))]
                        if st_type != "Cushion": 
                            fields.insert(1, ("Product", item.get('Product')))
                            if item.get('Arm/Table-Top'): fields.append(("Arm", item.get('Arm/Table-Top')))
                            if item.get('Panel'): fields.append(("Panel", item.get('Panel')))
                        if 'table' not in prod_l: fields.append(("Color", item.get('Color')))
                        
                        pdf.set_xy(cell_x, details_y); details_text = "".join([f"{l}: {v}\n" for l, v in fields if pd.notna(v) and str(v).strip()])
                        pdf.set_x(cell_x + (col_width-50)/2); pdf.set_text_color(100, 116, 139)
                        pdf.multi_cell(50, 3.5, details_text, ln=0, align='L')
                        current_col = (current_col + 1)
                        if current_col >= 3: current_col, current_row = 0, current_row + 1
                    
                    pdf_data = bytes(pdf.output())
                    st.download_button("Download PDF", data=pdf_data, file_name="NorthCape_Catalogue.pdf", mime="application/pdf", use_container_width=True)
                    
                    if st.button("Show Preview", use_container_width=True):
                        base64_pdf = base64.b64encode(pdf_data).decode('utf-8')
                        st.markdown(f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="800" type="application/pdf"></iframe>', unsafe_allow_html=True)
                except Exception as e: st.error(f"PDF Error: {str(e)}")

    # 3. Main Content Rendering
    st.markdown('<div class="hero-container"><div class="hero-title">NorthCape Image Library</div></div>', unsafe_allow_html=True)
    
    # Search Bar (Inside Fragment for speed)
    search_query = st.text_input("", placeholder="🔍 Search Part Number, Collection, Color...")
    if search_query:
        q = search_query.lower()
        searchable_cols = [c for c in filtered_df.columns if not any(x in c for x in ["Image", "Thumbnail", "Link", "List"])]
        mask = filtered_df[searchable_cols].apply(lambda row: row.astype(str).str.lower().str.contains(q).any(), axis=1)
        filtered_df = filtered_df[mask]

    st.caption(f"Showing {len(filtered_df)} records")
    
    # Pagination
    items_per_page = 25
    total_pages = max(1, (len(filtered_df) - 1) // items_per_page + 1)
    page_col1, page_col2 = st.columns([1, 4])
    with page_col1:
        page = st.number_input("Page", min_value=1, max_value=total_pages, value=1)
    
    start_idx = (page - 1) * items_per_page
    paged_data = filtered_df.iloc[start_idx:start_idx + items_per_page]

    # Grid Rendering
    grid_html = '<div class="card-grid">'
    TECHNICAL_FIELDS = ["Thumbnail", "Dropbox Folder Path", "Part Number", "Type", "Collection", "Collection Type", "Last Modified", "NC Image Count", "OS Image Count", "WF Image Count", "HD Image Count", "Local_Thumbnail", "Image_List", "Color_Link", "Part Number_Link"]

    for i, (_, item) in enumerate(paged_data.iterrows()):
        image_list = item.get("Image_List", []) or ([item["Local_Thumbnail"]] if item.get("Local_Thumbnail") else [])
        b64_images = [get_base64_img(t) for t in image_list[:3] if get_base64_img(t)]
        if not b64_images and item.get("Local_Thumbnail"):
            pb64 = get_base64_img(item.get("Local_Thumbnail"))
            if pb64: b64_images = [pb64]
            
        img_src = b64_images[0] if b64_images else ""
        is_shortlisted = item["Part Number"] in st.session_state.shortlist
        sc, si = ("active", "⭐") if is_shortlisted else ("", "☆")
        b64_data_attr = base64.b64encode(json.dumps(b64_images).encode()).decode()

        display_fields = []
        is_table = 'table' in str(item.get('Product', '')).lower()
        for k, v in item.items():
            if k not in TECHNICAL_FIELDS and not any(x in k for x in ["Image"]):
                if k == "Color" and is_table: continue
                if pd.notna(v) and str(v).strip() and str(v).lower() != 'nan':
                    display_fields.append((k, str(v)))

        def r_html(l, v):
            val = v
            if l == "Color" and pd.notna(item.get('Color_Link')):
                val = f'<a href="{item["Color_Link"]}" target="_blank" class="color-link">{v}</a>'
            return f'<div class="detail-row"><span class="detail-label">{l}</span><span class="detail-value">{val}</span></div>'

        stats_html = "".join([f'<div class="detail-row"><span class="detail-label">{k} Images</span><span class="detail-value">{int(item.get(c,0))}</span></div>' for k, c in [("NC", "NC Image Count"), ("OS", "OS Image Count"), ("WF", "WF Image Count"), ("HD", "HD Image Count")] if item.get(c,0)>0])
        swap_html = f'<div class="swap-btn" data-swap-target="img-{i}" title="Next Image" style="cursor: pointer; pointer-events: auto;">🔄</div>' if len(b64_images) > 1 else ""
        detail_rows = "".join([r_html(lbl, v) for lbl, v in display_fields])

        grid_html += (
            f'<div class="product-card">'
                f'<div class="shortlist-btn {sc}" data-part="{item["Part Number"]}" title="Add to Shortlist">{si}</div>'
                f'<div class="card-header"><div class="badge">{item["Collection Type"]}</div><div class="part-number">{item["Part Number"]}</div><div class="collection-text">{item["Collection"]}</div></div>'
                f'<div class="image-container"><img id="img-{i}" src="{img_src}" alt="Prod" data-urls-b64="{b64_data_attr}" data-idx="0">{swap_html}</div>'
                f'<div class="card-footer">{detail_rows}<div style="margin-top: 8px; border-top: 1px solid #f1f5f9; padding-top: 8px;">{stats_html}</div></div>'
            f'</div>'
        )

    st.markdown(grid_html + '</div>', unsafe_allow_html=True)

    # JS Injection (Bridge and Swapper)
    st.markdown("""
<script>
(function() {
    const p = window.parent.document;
    const h = (e) => {
        const b = e.target.closest('.swap-btn'); if (!b) return;
        e.preventDefault(); e.stopPropagation();
        const img = p.getElementById(b.getAttribute('data-swap-target')); if (!img) return;
        try {
            const urls = JSON.parse(atob(img.getAttribute('data-urls-b64')));
            let idx = (parseInt(img.getAttribute('data-idx')) || 0 + 1) % urls.length;
            img.src = urls[idx]; img.setAttribute('data-idx', idx);
        } catch(err) {}
    };
    const sh = (e) => {
        const b = e.target.closest('.shortlist-btn'); if (!b) return;
        const part = b.getAttribute('data-part');
        let inp = p.querySelector('input[placeholder="sync_bridge_v7"]');
        if (inp) {
            b.style.backgroundColor = '#fef08a'; b.style.transform = 'scale(0.8)';
            const set = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, "value").set;
            set.call(inp, part + "|" + Date.now());
            inp.dispatchEvent(new Event('input', { bubbles: true }));
            inp.dispatchEvent(new Event('change', { bubbles: true }));
            inp.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, keyCode: 13, key: 'Enter' }));
            setTimeout(() => { b.style.backgroundColor = ''; b.style.transform = ''; }, 300);
        }
    };
    p.removeEventListener('click', h); p.addEventListener('click', h);
    p.removeEventListener('click', sh); p.addEventListener('click', sh);
})();
</script>
""", unsafe_allow_html=True)


# Sidebar - Filtering
st.sidebar.title("")
selected_market = st.sidebar.selectbox("CHANNEL", ["Northcape", "Overstock", "Wayfair", "Home Depot"])

st.sidebar.divider()

# Cascaded Filter Logic (Turbo Speed)
def get_options(column, filtered_df):
    unique = filtered_df[column].unique().tolist()
    # Filter out nan, None, and empty strings strictly
    valid_options = [str(x) for x in unique if pd.notna(x) and x != "" and str(x).lower() != "nan"]
    return ["All"] + sorted(list(set(valid_options)))

df = pd.DataFrame(data)

# Channel to Count Column Mapping
channel_to_count = {
    "Northcape": "NC Image Count",
    "Overstock": "OS Image Count",
    "Wayfair": "WF Image Count",
    "Home Depot": "HD Image Count"
}

# Filter by Channel (Image Count > 0)
count_col = channel_to_count.get(selected_market)
if count_col and count_col in df.columns:
    # Force numeric conversion for reliability
    df[count_col] = pd.to_numeric(df[count_col], errors='coerce').fillna(0)
    df = df[df[count_col] > 0]

# Safety check for empty data or missing columns
if df.empty or "Collection Type" not in df.columns:
    st.error("⚠️ Catalogue data is missing or corrupted. Please run the update script.")
    st.sidebar.error("Data Load Error")
    if not df.empty:
        st.write("Columns found:", df.columns.tolist())
    st.stop()

# Filtering State
# Dynamic Type options based on data
type_options = get_options("Type", df)
selected_types = st.sidebar.multiselect("Type", type_options[1:], help="Select multiple product types") # Skip "All" for multiselect
filtered_df = df[df["Type"].isin(selected_types)] if selected_types else df

# The original 'Collection Type' contains the sheet/series names (2001, 6400, etc.)
series_options = get_options("Collection Type", filtered_df)
selected_series = st.sidebar.multiselect("Series", series_options[1:], help="Select multiple series")
if selected_series:
    filtered_df = filtered_df[filtered_df["Collection Type"].isin(selected_series)]

collection_options = get_options("Collection", filtered_df)
selected_collections = st.sidebar.multiselect("Collection", collection_options[1:], help="Select multiple collections")
if selected_collections:
    filtered_df = filtered_df[filtered_df["Collection"].isin(selected_collections)]

# Furniture specific filters
arm_opts = get_options("Arm/Table-Top", filtered_df)
if len(arm_opts) > 1:
    selected_arms = st.sidebar.multiselect("Arm/Table-Top", arm_opts[1:])
    if selected_arms:
        filtered_df = filtered_df[filtered_df["Arm/Table-Top"].isin(selected_arms)]

product_opts = get_options("Product", filtered_df)
if len(product_opts) > 1:
    selected_products = st.sidebar.multiselect("Product", product_opts[1:])
    if selected_products:
        filtered_df = filtered_df[filtered_df["Product"].isin(selected_products)]

panel_opts = get_options("Panel", filtered_df)
if len(panel_opts) > 1:
    selected_panels = st.sidebar.multiselect("Panel", panel_opts[1:])
    if selected_panels:
        filtered_df = filtered_df[filtered_df["Panel"].isin(selected_panels)]

color_opts = get_options("Color", filtered_df)
selected_colors = st.sidebar.multiselect("Color", color_opts[1:])
if selected_colors:
    filtered_df = filtered_df[filtered_df["Color"].isin(selected_colors)]

# Run the fragment
render_main_ui(filtered_df)
