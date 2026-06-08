"""
update_from_jun05_excel.py
Sync catalogue.json with Jun 05, 2026 Excel files.
"""

import csv
import hashlib
import io
import json
import os
import re
import shutil
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from PIL import Image

# Force line-buffered stdout so progress prints appear immediately
sys.stdout.reconfigure(line_buffering=True)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path("c:/Users/Lenovo/Downloads/NC - Library Naming/Viewer/NC_Turbo_Viewer")
CATALOGUE_PATH = BASE_DIR / "data" / "catalogue.json"
THUMB_DIR = BASE_DIR / "static" / "thumbnails"

EXCEL_FILES = [
    "C:/Users/Lenovo/Master Image Library Dropbox/Master Image Library/NorthCape Library - Master Excel - Accent Pillow - Jun 05, 2026.xlsx",
    "C:/Users/Lenovo/Master Image Library Dropbox/Master Image Library/NorthCape Library - Master Excel - Cushions - Jun 05, 2026.xlsx",
    "C:/Users/Lenovo/Master Image Library Dropbox/Master Image Library/NorthCape Library - Master Excel - Furniture - Jun 05, 2026.xlsx",
]

CSV_REPLACEMENTS = "C:/Users/Lenovo/Desktop/To be added to Library/replacements_to_delete.csv"

LOG_PATH = BASE_DIR / "update_jun05_log.csv"

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SKIP_SHEETS = {"6600 - DG", "6701 - DG"}
SKIP_PN_PREFIXES = ("NC6600", "NCI6600", "NC6701", "NCI6701")

IMAGE_COLUMN_ORDER = (
    ["Wayfair Image {}".format(i) for i in range(1, 16)]
    + ["Overstock Image {}".format(i) for i in range(1, 16)]
    + ["Home Depot Image {}".format(i) for i in range(1, 16)]
    + ["Northcape Image {}".format(i) for i in range(1, 16)]
    + ["Bed Bath Beyond Image {}".format(i) for i in range(1, 16)]
)
MAX_IMAGES = 5
THUMB_SIZE = (400, 400)
THUMB_QUALITY = 85
DOWNLOAD_TIMEOUT = 25
DOWNLOAD_WORKERS = 16  # parallel download threads


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def url_to_thumb_name(url: str) -> str:
    return hashlib.md5(url.encode()).hexdigest() + ".jpg"


def thumb_rel(name: str) -> str:
    return f"thumbnails/{name}"


def normalize_dropbox_url(url: str) -> str:
    """Replace dl=0/dl=1 with raw=1 for direct download."""
    url = re.sub(r'[?&]dl=\d', '', url)
    url = re.sub(r'[?&]raw=\d', '', url)
    if '?' in url:
        url += '&raw=1'
    else:
        url += '?raw=1'
    return url


def is_skipped_pn(pn: str) -> bool:
    return pn.startswith(SKIP_PN_PREFIXES)


def nci_to_nc_suffix(pn: str):
    """If pn starts with NCI, return the suffix (after NCI). Else None."""
    if pn.startswith("NCI"):
        return pn[3:]
    return None


def nc_suffix(pn: str):
    """If pn starts with NC (not NCI), return suffix after NC. Else None."""
    if pn.startswith("NC") and not pn.startswith("NCI"):
        return pn[2:]
    return None


# ---------------------------------------------------------------------------
# Step 1: Read Excel files
# ---------------------------------------------------------------------------

def read_excel_products() -> dict:
    """Return dict keyed by Part Number with all product fields."""
    products = {}

    for excel_path in EXCEL_FILES:
        print(f"  Reading: {Path(excel_path).name}")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                print(f"    Skipping sheet: {sheet_name}")
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row (first row containing 'Part Number')
            header = None
            header_row_idx = None
            for idx, row in enumerate(rows):
                row_strs = [str(c or "").strip() for c in row]
                if "Part Number" in row_strs:
                    header = row_strs
                    header_row_idx = idx
                    break

            if header is None:
                print(f"    WARNING: No header row in {sheet_name}, skipping")
                continue

            pn_idx = header.index("Part Number")
            col_map = {name: i for i, name in enumerate(header) if name}

            def get_col(row, col_name, default=""):
                i = col_map.get(col_name)
                if i is None or i >= len(row):
                    return default
                return str(row[i] or "").strip()

            # Color column varies by sheet
            color_col = next((c for c in ("Color", "Cushion Color") if c in col_map), None)
            # Category column
            category_col = next((c for c in ("Category", "Type") if c in col_map), None)
            type_col = "Type" if "Type" in col_map else None

            count = 0
            for row in rows[header_row_idx + 1:]:
                if not row or row[pn_idx] is None:
                    continue
                pn = str(row[pn_idx] or "").strip()
                if not pn or is_skipped_pn(pn):
                    continue

                # Collect first 5 image URLs in priority order
                image_urls = []
                all_image_cols = {}
                for col_name in IMAGE_COLUMN_ORDER:
                    i = col_map.get(col_name)
                    if i is None or i >= len(row):
                        all_image_cols[col_name] = ""
                    else:
                        val = str(row[i] or "").strip()
                        all_image_cols[col_name] = val
                        if val and len(image_urls) < MAX_IMAGES:
                            image_urls.append(val)

                color = get_col(row, color_col) if color_col else ""
                category = get_col(row, category_col) if category_col else ""
                prod_type = get_col(row, type_col) if type_col else ""

                products[pn] = {
                    "Part Number": pn,
                    "Collection": get_col(row, "Collection"),
                    "Color": color,
                    "Category": category,
                    "Type": prod_type,
                    "Dropbox Folder Path": get_col(row, "Dropbox Folder Path"),
                    "Last Modified": get_col(row, "Last Modified"),
                    "NC Image Count": get_col(row, "NC Image Count", "0"),
                    "BY Image Count": get_col(row, "BY Image Count", "0"),
                    "WF Image Count": get_col(row, "WF Image Count", "0"),
                    "HD Image Count": get_col(row, "HD Image Count", "0"),
                    "Thumbnail": get_col(row, "Thumbnail"),
                    "_thumbnail_path": get_col(row, "_thumbnail_path"),
                    "Color_Link": get_col(row, "Color_Link"),
                    "_image_urls": image_urls,
                    "_all_image_cols": all_image_cols,
                    "_sheet": sheet_name,
                    "_excel": Path(excel_path).stem,
                }
                count += 1

            print(f"    Sheet '{sheet_name}': {count} products")

        wb.close()

    print(f"Total Excel products (ex 6600/6701): {len(products)}")
    return products


# ---------------------------------------------------------------------------
# Step 2: Build change plan
# ---------------------------------------------------------------------------

def get_catalogue_image_urls(cat_product: dict) -> list:
    """Extract the ordered first-5 URLs from a catalogue product."""
    urls = []
    for col_name in IMAGE_COLUMN_ORDER:
        if col_name in cat_product:
            val = str(cat_product[col_name] or "").strip()
            if val and len(urls) < MAX_IMAGES:
                urls.append(val)
    return urls


def build_change_plan(excel_products: dict, catalogue: list) -> dict:
    cat_by_pn = {p["Part Number"]: p for p in catalogue}

    # Build NC-suffix reverse map for catalogue NC-prefixed products
    nc_suffix_map = {}
    for pn in cat_by_pn:
        sfx = nc_suffix(pn)
        if sfx is not None:
            nc_suffix_map[sfx] = pn

    renames = []
    url_changes = []
    new_products = []
    unchanged_pns = []
    accounted_cat_pns = set()

    # Never touch 6600/6701 products in catalogue
    skipped_cat_pns = {p["Part Number"] for p in catalogue if is_skipped_pn(p["Part Number"])}
    accounted_cat_pns.update(skipped_cat_pns)

    for excel_pn, excel_prod in excel_products.items():
        excel_urls = excel_prod["_image_urls"]

        # Direct match
        if excel_pn in cat_by_pn:
            cat_prod = cat_by_pn[excel_pn]
            cat_urls = get_catalogue_image_urls(cat_prod)
            accounted_cat_pns.add(excel_pn)
            if excel_urls == cat_urls:
                unchanged_pns.append(excel_pn)
            else:
                url_changes.append((excel_pn, cat_prod, excel_prod, cat_urls, excel_urls))
            continue

        # NC->NCI rename check
        sfx = nci_to_nc_suffix(excel_pn)
        if sfx is not None and sfx in nc_suffix_map:
            old_cat_pn = nc_suffix_map[sfx]
            cat_prod = cat_by_pn[old_cat_pn]
            accounted_cat_pns.add(old_cat_pn)
            renames.append((old_cat_pn, excel_pn, cat_prod, excel_prod))
            cat_urls = get_catalogue_image_urls(cat_prod)
            if excel_urls != cat_urls:
                url_changes.append((excel_pn, cat_prod, excel_prod, cat_urls, excel_urls))
            continue

        # New product
        new_products.append(excel_prod)

    removed_pns = [pn for pn in cat_by_pn if pn not in accounted_cat_pns]

    print(f"Change plan:")
    print(f"  Renames (NC->NCI):   {len(renames)}")
    print(f"  URL changes:         {len(url_changes)}")
    print(f"  New products:        {len(new_products)}")
    print(f"  Removed products:    {len(removed_pns)}")
    print(f"  Unchanged:           {len(unchanged_pns)}")

    return {
        "renames": renames,
        "url_changes": url_changes,
        "new_products": new_products,
        "removed_pns": removed_pns,
        "unchanged_pns": unchanged_pns,
    }


# ---------------------------------------------------------------------------
# Step 3: Handle replacements_to_delete.csv
# ---------------------------------------------------------------------------

def read_replacements() -> set:
    """Return set of Part Numbers whose thumbnails must be force-refreshed."""
    pns = set()
    with open(CSV_REPLACEMENTS, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pn = row.get("Raw Part Number", "").strip()
            if pn and not is_skipped_pn(pn):
                pns.add(pn)
    print(f"Replacements CSV: {len(pns)} unique Part Numbers to force-refresh")
    return pns


# ---------------------------------------------------------------------------
# Step 4: Execute
# ---------------------------------------------------------------------------

def download_one(url: str, dest_path: Path) -> tuple:
    """Download and resize one thumbnail. Returns (url, tname, ok, err_msg)."""
    tname = dest_path.name
    try:
        dl_url = normalize_dropbox_url(url)
        resp = requests.get(dl_url, timeout=DOWNLOAD_TIMEOUT)
        resp.raise_for_status()
        img = Image.open(io.BytesIO(resp.content)).convert("RGB")
        img = img.resize(THUMB_SIZE, Image.LANCZOS)
        img.save(dest_path, "JPEG", quality=THUMB_QUALITY)
        return (url, tname, True, "")
    except Exception as e:
        return (url, tname, False, str(e)[:120])


def run_update(plan: dict, catalogue: list, excel_products: dict, replacement_pns: set):
    cat_by_pn = {p["Part Number"]: p for p in catalogue}

    thumbs_to_delete: set = set()   # thumb filenames (basename only)
    # dl_queue: list of (url, tname) — deduplicated later
    dl_queue_raw: list = []
    log_rows = []

    # ---- Apply renames ----
    print(f"  Applying {len(plan['renames'])} renames...")
    for old_pn, new_pn, cat_prod, excel_prod in plan["renames"]:
        cat_prod["Part Number"] = new_pn
        cat_by_pn[new_pn] = cat_prod
        del cat_by_pn[old_pn]
        log_rows.append({
            "Part Number": new_pn,
            "Collection": cat_prod.get("Collection", ""),
            "Color": cat_prod.get("Color", ""),
            "Action": "RENAME",
            "Details": f"NC->NCI: {old_pn} -> {new_pn}",
            "Result": "OK",
        })

    # ---- Apply URL changes ----
    print(f"  Applying {len(plan['url_changes'])} URL changes...")
    for pn, cat_prod, excel_prod, old_urls, new_urls in plan["url_changes"]:
        effective_pn = cat_prod["Part Number"]
        old_thumbs = {url_to_thumb_name(u) for u in old_urls if u}
        new_thumbs = {url_to_thumb_name(u) for u in new_urls if u}
        to_del = old_thumbs - new_thumbs
        thumbs_to_delete.update(to_del)

        for url in new_urls:
            if url:
                dl_queue_raw.append((url, url_to_thumb_name(url)))

        _apply_image_fields(cat_prod, excel_prod, new_urls)
        log_rows.append({
            "Part Number": effective_pn,
            "Collection": cat_prod.get("Collection", ""),
            "Color": cat_prod.get("Color", ""),
            "Action": "URL_CHANGE",
            "Details": f"old={len(old_urls)} new={len(new_urls)} del_thumbs={len(to_del)}",
            "Result": "OK",
        })

    # ---- Force-refresh (replacements) ----
    print(f"  Applying {len(replacement_pns)} force-refreshes...")
    for pn in replacement_pns:
        if pn not in cat_by_pn:
            log_rows.append({
                "Part Number": pn,
                "Collection": "", "Color": "",
                "Action": "REPLACEMENT_SKIP",
                "Details": "Not found in catalogue",
                "Result": "SKIP",
            })
            continue
        cat_prod = cat_by_pn[pn]
        cat_urls = get_catalogue_image_urls(cat_prod)
        for url in cat_urls:
            if url:
                tname = url_to_thumb_name(url)
                thumbs_to_delete.add(tname)   # delete first
                dl_queue_raw.append((url, tname))  # then re-download

        log_rows.append({
            "Part Number": pn,
            "Collection": cat_prod.get("Collection", ""),
            "Color": cat_prod.get("Color", ""),
            "Action": "FORCE_REFRESH",
            "Details": f"Force re-download {len(cat_urls)} thumbnails",
            "Result": "OK",
        })

    # ---- New products ----
    print(f"  Adding {len(plan['new_products'])} new products...")
    for excel_prod in plan["new_products"]:
        pn = excel_prod["Part Number"]
        new_urls = excel_prod["_image_urls"]
        new_cat_prod = _build_new_cat_product(excel_prod, new_urls)
        catalogue.append(new_cat_prod)
        cat_by_pn[pn] = new_cat_prod
        for url in new_urls:
            if url:
                dl_queue_raw.append((url, url_to_thumb_name(url)))
        log_rows.append({
            "Part Number": pn,
            "Collection": excel_prod.get("Collection", ""),
            "Color": excel_prod.get("Color", ""),
            "Action": "NEW",
            "Details": f"new product, {len(new_urls)} images",
            "Result": "OK",
        })

    # ---- Remove products ----
    print(f"  Removing {len(plan['removed_pns'])} products...")
    removed_set = set(plan["removed_pns"])
    for pn in plan["removed_pns"]:
        cat_prod = cat_by_pn.get(pn)
        if cat_prod:
            cat_urls = get_catalogue_image_urls(cat_prod)
            for url in cat_urls:
                if url:
                    thumbs_to_delete.add(url_to_thumb_name(url))
            log_rows.append({
                "Part Number": pn,
                "Collection": cat_prod.get("Collection", ""),
                "Color": cat_prod.get("Color", ""),
                "Action": "REMOVED",
                "Details": f"Removed, {len(cat_urls)} thumbs queued for deletion",
                "Result": "OK",
            })

    # ---- Update metadata for unchanged products ----
    for pn in plan["unchanged_pns"]:
        if pn in cat_by_pn and pn in excel_products:
            _update_metadata_only(cat_by_pn[pn], excel_products[pn])

    # ---- Remove deleted products from catalogue ----
    catalogue[:] = [p for p in catalogue if p["Part Number"] not in removed_set]

    # ---- Build final thumbnail reference set ----
    final_thumb_refs: set = set()
    for p in catalogue:
        for rel_path in p.get("Image_List", []):
            final_thumb_refs.add(Path(rel_path).name)

    # ---- 4a: Delete old thumbnails ----
    # For force-refresh: delete BEFORE checking "already exists" in download
    safe_to_delete = thumbs_to_delete - final_thumb_refs
    # Force-refresh thumbnails were added to thumbs_to_delete; delete them even if referenced
    # (they'll be re-added after download). But we need to delete them before download step.
    # So delete ALL in thumbs_to_delete first, then filter at download step by existence.
    force_refresh_thumbs = set()
    for pn in replacement_pns:
        if pn in cat_by_pn:
            cat_prod = cat_by_pn[pn]
            cat_urls = get_catalogue_image_urls(cat_prod)
            for url in cat_urls:
                if url:
                    force_refresh_thumbs.add(url_to_thumb_name(url))

    # Delete: safe_to_delete (not referenced by any product after update) + force_refresh
    all_to_delete = safe_to_delete | force_refresh_thumbs
    deleted_count = 0
    for tname in all_to_delete:
        path = THUMB_DIR / tname
        if path.exists():
            path.unlink()
            deleted_count += 1

    print(f"Deleted {deleted_count} thumbnails ({len(thumbs_to_delete)} queued, {len(thumbs_to_delete & final_thumb_refs) - len(force_refresh_thumbs)} skipped/still-referenced)")

    # ---- 4b: Parallel download ----
    # Deduplicate
    seen_dl: set = set()
    dl_queue = []
    for url, tname in dl_queue_raw:
        if tname not in seen_dl:
            seen_dl.add(tname)
            dl_queue.append((url, tname))

    # Only download files not already on disk
    actually_needed = [(url, tname) for url, tname in dl_queue if not (THUMB_DIR / tname).exists()]
    print(f"Download queue: {len(dl_queue)} unique, {len(dl_queue) - len(actually_needed)} already on disk, {len(actually_needed)} to fetch")

    downloaded = 0
    failed = 0
    failures = []
    def do_download(args):
        url, tname = args
        dest = THUMB_DIR / tname
        return download_one(url, dest)

    with ThreadPoolExecutor(max_workers=DOWNLOAD_WORKERS) as executor:
        futures = {executor.submit(do_download, item): item for item in actually_needed}
        for future in as_completed(futures):
            url, tname, ok, err = future.result()
            if ok:
                downloaded += 1
            else:
                failed += 1
                failures.append((tname, err))
            done = downloaded + failed
            if done % 50 == 0 or done == len(actually_needed):
                print(f"  Progress: {done}/{len(actually_needed)} ({failed} failures)")

    if failures:
        print(f"  Failed downloads ({len(failures)}):")
        for tname, err in failures[:20]:
            print(f"    {tname}: {err}")

    print(f"Downloads complete: {downloaded} ok, {failed} failed")

    # ---- 4c: Rebuild Image_List/Local_Thumbnail for all updated products ----
    for p in catalogue:
        pn = p["Part Number"]
        if pn not in excel_products:
            continue
        excel_prod = excel_products[pn]
        new_urls = excel_prod["_image_urls"]
        img_list = []
        for url in new_urls:
            if url:
                tname = url_to_thumb_name(url)
                if (THUMB_DIR / tname).exists():
                    img_list.append(thumb_rel(tname))
        if img_list:
            p["Image_List"] = img_list
            p["Local_Thumbnail"] = img_list[0]

    return {
        "log_rows": log_rows,
        "deleted_count": deleted_count,
        "downloaded": downloaded,
        "failed": failed,
        "dl_queue_size": len(actually_needed),
    }


def _apply_image_fields(cat_prod: dict, excel_prod: dict, new_urls: list):
    """Update image URL fields and metadata in cat_prod from excel_prod."""
    for col_name in IMAGE_COLUMN_ORDER:
        if col_name in excel_prod.get("_all_image_cols", {}):
            cat_prod[col_name] = excel_prod["_all_image_cols"][col_name]

    new_img_list = [thumb_rel(url_to_thumb_name(u)) for u in new_urls if u]
    cat_prod["Image_List"] = new_img_list
    cat_prod["Local_Thumbnail"] = new_img_list[0] if new_img_list else ""

    for field in ("Collection", "Color", "Category", "Type", "Dropbox Folder Path",
                  "Last Modified", "NC Image Count", "BY Image Count",
                  "WF Image Count", "HD Image Count"):
        if field in excel_prod and excel_prod[field] not in ("", None):
            cat_prod[field] = excel_prod[field]


def _update_metadata_only(cat_prod: dict, excel_prod: dict):
    """Update non-image metadata for unchanged products."""
    for field in ("Collection", "Color", "Category", "Type", "Dropbox Folder Path",
                  "Last Modified", "NC Image Count", "BY Image Count",
                  "WF Image Count", "HD Image Count"):
        if field in excel_prod and excel_prod[field] not in ("", None):
            cat_prod[field] = excel_prod[field]


def _build_new_cat_product(excel_prod: dict, new_urls: list) -> dict:
    """Build a new catalogue product record from Excel data."""
    img_list = [thumb_rel(url_to_thumb_name(u)) for u in new_urls if u]
    prod = {
        "Part Number": excel_prod["Part Number"],
        "Collection": excel_prod.get("Collection", ""),
        "Color": excel_prod.get("Color", ""),
        "Category": excel_prod.get("Category", ""),
        "Type": excel_prod.get("Type", ""),
        "Dropbox Folder Path": excel_prod.get("Dropbox Folder Path", ""),
        "Last Modified": excel_prod.get("Last Modified", ""),
        "NC Image Count": excel_prod.get("NC Image Count", 0),
        "BY Image Count": excel_prod.get("BY Image Count", 0),
        "WF Image Count": excel_prod.get("WF Image Count", 0),
        "HD Image Count": excel_prod.get("HD Image Count", 0),
        "Thumbnail": excel_prod.get("Thumbnail", ""),
        "_thumbnail_path": excel_prod.get("_thumbnail_path", ""),
        "Color_Link": excel_prod.get("Color_Link", ""),
        "Collection Type": "",
        "Image_List": img_list,
        "Local_Thumbnail": img_list[0] if img_list else "",
    }
    for col_name in IMAGE_COLUMN_ORDER:
        prod[col_name] = excel_prod.get("_all_image_cols", {}).get(col_name, "")
    return prod


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    start = time.time()
    print("=" * 60)
    print("update_from_jun05_excel.py")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Backup
    backup_path = CATALOGUE_PATH.parent / f"catalogue.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    shutil.copy2(CATALOGUE_PATH, backup_path)
    print(f"Backup saved: {backup_path.name}")

    # Step 1
    print("\n[Step 1] Reading Excel files...")
    excel_products = read_excel_products()

    print("\nLoading catalogue.json...")
    with open(CATALOGUE_PATH, encoding="utf-8") as f:
        catalogue = json.load(f)
    print(f"Catalogue: {len(catalogue)} products")

    # Step 2
    print("\n[Step 2] Building change plan...")
    plan = build_change_plan(excel_products, catalogue)

    # Step 3
    print("\n[Step 3] Reading replacements CSV...")
    replacement_pns = read_replacements()

    # Step 4
    print("\n[Step 4] Executing changes...")
    result = run_update(plan, catalogue, excel_products, replacement_pns)

    # Step 5
    print("\n[Step 5] Saving catalogue.json...")
    with open(CATALOGUE_PATH, "w", encoding="utf-8") as f:
        json.dump(catalogue, f, ensure_ascii=False, indent=2)
    print(f"Saved {len(catalogue)} products")

    log_rows = result["log_rows"]
    with open(LOG_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Part Number", "Collection", "Color", "Action", "Details", "Result"])
        writer.writeheader()
        writer.writerows(log_rows)
    print(f"Log written: {LOG_PATH.name} ({len(log_rows)} rows)")

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  Products renamed (NC->NCI):  {len(plan['renames'])}")
    print(f"  Products with URL changes:   {len(plan['url_changes'])}")
    print(f"  New products added:          {len(plan['new_products'])}")
    print(f"  Products removed:            {len(plan['removed_pns'])}")
    print(f"  Thumbnails deleted:          {result['deleted_count']}")
    print(f"  Thumbnails downloaded:       {result['downloaded']}")
    print(f"  Download failures:           {result['failed']}")
    print(f"  Total catalogue size:        {len(catalogue)}")
    print(f"  Elapsed: {time.time() - start:.1f}s")
    print("=" * 60)


if __name__ == "__main__":
    main()
