import openpyxl
import json
import re
import os
import requests
from PIL import Image
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor
import hashlib
import shutil
import glob

# --- Configuration ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
THUMBNAIL_SIZE = (400, 400)
MAX_THREADS = 15  # Slower but more reliable (prevents throttling)
PUBLIC_THUMBS_DIR = os.path.join(BASE_DIR, "static", "thumbnails")
DATA_DIR = os.path.join(BASE_DIR, "data")
# Support multiple excel files - but use only the NEWEST cushion file to avoid duplicates
# Sort by modification time descending so newest come first
all_excel = glob.glob(os.path.join(BASE_DIR, "*.xlsx"))
all_excel.sort(key=lambda p: os.path.getmtime(p), reverse=True)

# De-duplicate: if multiple files have the same sheet names, keep only the newest
# For cushion files specifically, keep only the newest one
EXCEL_FILES = []
seen_cushion = False
for f in all_excel:
    fname_lower = os.path.basename(f).lower()
    # Skip temp/lock files
    if os.path.basename(f).startswith('~'):
        continue
    # For cushion-specific files, only keep the newest
    if 'cushion' in fname_lower:
        if not seen_cushion:
            EXCEL_FILES.append(f)
            seen_cushion = True
        # else skip older cushion files
    else:
        EXCEL_FILES.append(f)

JSON_OUTPUT = os.path.join(DATA_DIR, "catalogue.json")

def extract_link(formula):
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    match = re.search(r'HYPERLINK\("(.*?)",\s*"(.*?)"\)', formula)
    if match:
        return match.group(1)
    return None

def get_value_from_formula(formula):
    if not isinstance(formula, str) or not formula.startswith("="):
        return str(formula) if formula is not None else ""
    match = re.search(r'HYPERLINK\("(.*?)",\s*"(.*?)"\)', formula)
    if match:
        return match.group(2)
    return str(formula)

def get_legacy_urls(url):
    """Generate old-naming URL variants to find existing cached thumbnails.

    Handles the rename: NC->NCI (Northcape), WF-NC->NC (Wayfair), BY- unchanged.
    Given a NEW url, produces the OLD url so we can find cached thumbnails.
    """
    legacy = []
    match = re.search(r'(/fi/[^/]+/)([^?]+)', url)
    if not match:
        return legacy
    path_prefix, filename = match.group(1), match.group(2)

    # NCI -> NC (Northcape was renamed from NC to NCI)
    if re.match(r'NCI\d', filename):
        old_name = re.sub(r'^NCI', 'NC', filename)
        legacy.append(url.replace(path_prefix + filename, path_prefix + old_name))

    # NC (standalone) -> WF-NC (Wayfair had WF- prefix before)
    if re.match(r'NC\d', filename) and not filename.startswith('BY-'):
        old_name = 'WF-' + filename
        legacy.append(url.replace(path_prefix + filename, path_prefix + old_name))

    return legacy

def download_and_resize(url, part_number):
    if not url or "dropbox.com" not in url:
        return None
    
    url_hash = hashlib.md5(url.encode()).hexdigest()
    thumb_name = f"{url_hash}.jpg"
    thumb_path = os.path.join(PUBLIC_THUMBS_DIR, thumb_name)
    
    # Skip if already exists (avoids duplicate work)
    if os.path.exists(thumb_path):
        return f"thumbnails/{thumb_name}"

    # Check for thumbnail cached under legacy (old naming) URL
    for legacy_url in get_legacy_urls(url):
        legacy_hash = hashlib.md5(legacy_url.encode()).hexdigest()
        legacy_path = os.path.join(PUBLIC_THUMBS_DIR, f"{legacy_hash}.jpg")
        if os.path.exists(legacy_path):
            shutil.copy2(legacy_path, thumb_path)
            return f"thumbnails/{thumb_name}"

    try:
        raw_url = url.replace("dl=0", "raw=1").replace("dl=1", "raw=1")
        # Increased timeout for reliability
        response = requests.get(raw_url, timeout=30)
        if response.status_code == 200:
            img = Image.open(BytesIO(response.content))
            if img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
            img.thumbnail(THUMBNAIL_SIZE)
            img.save(thumb_path, "JPEG", quality=85)
            # Return relative path for the app
            return f"thumbnails/{thumb_name}"
    except Exception as e:
        pass
    return None

def update_catalogue():
    print("--- NorthCape Turbo Upkeep Script v2 ---")
    if not EXCEL_FILES:
        print(f"Error: No Excel files found in {BASE_DIR}!")
        return

    os.makedirs(PUBLIC_THUMBS_DIR, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)
    
    catalogue_data = []

    for excel_path in EXCEL_FILES:
        print(f"Loading Excel: {os.path.basename(excel_path)}...")
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ["Status", "Empty", "Sheet1"]: continue
            
            print(f"  Processing Sheet: {sheet_name}")
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1] if cell.value]
            
            for row_idx in range(2, ws.max_row + 1):
                row_data = {"Collection Type": sheet_name}
                has_data = False
                
                for col_idx, header in enumerate(headers):
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    val = cell.value
                    
                    # Normalize Headers (Cushion Color -> Color)
                    norm_header = header
                    if header == "Cushion Color": norm_header = "Color"
                    
                    if norm_header in ["Color", "Part Number"]:
                        row_data[norm_header] = get_value_from_formula(val)
                        link = extract_link(val)
                        if link: row_data[f"{norm_header}_Link"] = link
                    elif norm_header == "Type":
                        # Normalize Cushions (plural) to Cushion (singular)
                        val_str = str(val) if val else ""
                        if val_str.lower() == "cushions":
                            row_data[norm_header] = "Cushion"
                        else:
                            row_data[norm_header] = val_str
                    else:
                        row_data[norm_header] = val if val is not None else ""
                    
                    if val: has_data = True
                
                # Collect ALL potential Dropbox links for this item
                potential_urls = []
                img_cols = sorted([h for h in headers if h and "Image" in h and h != "Dropbox Folder Path"], 
                                 key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 999)
                
                for h in img_cols:
                    val = row_data.get(h)
                    if val and "dropbox.com" in str(val):
                        potential_urls.append(str(val))
                
                if has_data:
                    row_data["_potential_urls"] = potential_urls
                    catalogue_data.append(row_data)

    stats = {"primary_1": 0, "no_logo": 0, "fallback": 0}
    
    def process_item(item):
        urls = item.get("_potential_urls", [])
        blacklist = ["logo", "sunbrella", "branding", "infographic", "text", "warranty"]
        
        # Download and collect pairs of (original_url, local_thumb_path)
        url_thumb_pairs = []
        for url in urls:
            thumb_path = download_and_resize(url, item.get("Part Number"))
            if thumb_path:
                url_thumb_pairs.append((url, thumb_path))
                if len(url_thumb_pairs) >= 5: break # Cap at 5 images
        
        item["Image_List"] = [p[1] for p in url_thumb_pairs]
        item["Local_Thumbnail"] = None
        
        if url_thumb_pairs:
            # Ranking Algorithm for Primary Thumbnail:
            # 1. Best: URL contains "_1." (standard primary shot) and is NOT a logo/blacklist
            for url, thumb in url_thumb_pairs:
                url_lower = url.lower()
                is_manual_primary = "_1." in url_lower or "_1_" in url_lower
                is_blacklisted = any(word in url_lower for word in blacklist)
                
                if is_manual_primary and not is_blacklisted:
                    item["Local_Thumbnail"] = thumb
                    stats["primary_1"] += 1
                    break
            
            # 2. Second Best: Any image that is NOT a logo/blacklist
            if not item["Local_Thumbnail"]:
                for url, thumb in url_thumb_pairs:
                    if not any(word in url.lower() for word in blacklist):
                        item["Local_Thumbnail"] = thumb
                        stats["no_logo"] += 1
                        break
            
            # 3. Fallback: Just use the very first available image
            if not item["Local_Thumbnail"]:
                item["Local_Thumbnail"] = url_thumb_pairs[0][1]
                stats["fallback"] += 1
                
        if "_potential_urls" in item:
            del item["_potential_urls"]
        return item

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        final_data = list(executor.map(process_item, catalogue_data))

    print("\n--- Thumbnail Statistics ---")
    print(f"  [OK] Priority _1 Image: {stats['primary_1']}")
    print(f"  [i]  Clean product shot (no logo): {stats['no_logo']}")
    print(f"  [!]  Fallback (may include logo): {stats['fallback']}")
    print("---------------------------\n")

    with open(JSON_OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(final_data, f, indent=2, ensure_ascii=False)
    
    print(f"Success! {len(final_data)} records updated. App is ready.")

if __name__ == "__main__":
    update_catalogue()
